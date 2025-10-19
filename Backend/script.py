from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import numpy as np
from pyngrok import ngrok
import os
from dotenv import load_dotenv
from inference_sdk import InferenceHTTPClient
import tempfile
from PIL import Image
from datetime import datetime
import requests
import json
import pandas as pd
import openpyxl

# ML integration -----------------------------------------------------------
# Import ML helpers from ML.py. This allows app.py to call ML.predict(...) without
# retraining; artifacts are loaded from ml_artifacts.joblib by load_artifacts().
from ML import load_artifacts, predict

# Load model artifacts at startup (best-effort). If artifacts are missing the app
# will still start; run `python ML.py` once to create artifacts (ml_artifacts.joblib).
try:
    load_artifacts()
except Exception:
    # Artifacts not present yet — that's okay for now. They can be created by
    # running ML.py manually. We avoid failing fast so development can continue.
    pass
# -------------------------------------------------------------------------

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app, origins="*", methods=["GET", "POST", "OPTIONS"], allow_headers=["Content-Type", "Authorization"])

# Initialize Roboflow client
client = InferenceHTTPClient(
    api_url="https://serverless.roboflow.com",
    api_key=os.getenv("ROBOFLOW_API_KEY")
)

# Create photos folder if it doesn't exist
PHOTOS_FOLDER = os.path.join(os.path.dirname(__file__), 'photos')
if not os.path.exists(PHOTOS_FOLDER):
    os.makedirs(PHOTOS_FOLDER)
    print(f"Created photos folder: {PHOTOS_FOLDER}")

# Excel file path for tracking
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'image_location_log.xlsx')

def get_location_from_wifi(wifi_data):
    """Get location using Google Geolocation API with WiFi data"""
    try:
        api_key = os.getenv("GOOGLE_GEOLOCATION_API_KEY")
        if not api_key:
            return None, None, "No Google API key"
        
        url = f"https://www.googleapis.com/geolocation/v1/geolocate?key={api_key}"
        
        # Parse WiFi data from ESP32
        wifi_json = json.loads(wifi_data)
        
        response = requests.post(url, json=wifi_json, timeout=10)
        if response.status_code == 200:
            location_data = response.json()
            lat = location_data.get('location', {}).get('lat')
            lng = location_data.get('location', {}).get('lng')
            accuracy = location_data.get('accuracy', 'unknown')
            return lat, lng, f"accuracy: {accuracy}m"
        else:
            return None, None, f"API error: {response.status_code}"
    except Exception as e:
        return None, None, f"Error: {str(e)}"

def log_to_excel(image_name, latitude, longitude, location_info, trash_detected, trash_type, timestamp):
    """Log image information to Excel file"""
    try:
        # Create new row data
        new_row = {
            'Timestamp': timestamp,
            'Image_Name': image_name,
            'Latitude': latitude if latitude else 'N/A',
            'Longitude': longitude if longitude else 'N/A',
            'Location_Info': location_info,
            'Trash_Detected': trash_detected,
            'Trash_Type': trash_type
        }
        
        # Check if Excel file exists
        if os.path.exists(EXCEL_FILE):
            # Read existing data
            df = pd.read_excel(EXCEL_FILE)
            # Append new row
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        else:
            # Create new DataFrame
            df = pd.DataFrame([new_row])
        
        # Save to Excel
        df.to_excel(EXCEL_FILE, index=False)
        print(f"Logged to Excel: {image_name}")
        
    except Exception as e:
        print(f"Error logging to Excel: {str(e)}")

def rgb565_to_jpeg(rgb565_data, width=320, height=240):
    """Convert RGB565 data to JPEG format - QVGA resolution"""
    # Convert RGB565 bytes to numpy array
    rgb565_array = np.frombuffer(rgb565_data, dtype=np.uint16)
    
    # Convert RGB565 to RGB888
    r = ((rgb565_array & 0xF800) >> 11) << 3
    g = ((rgb565_array & 0x07E0) >> 5) << 2
    b = (rgb565_array & 0x001F) << 3
    
    # Create RGB array
    rgb_array = np.zeros((height, width, 3), dtype=np.uint8)
    rgb_array[:, :, 0] = r.reshape(height, width)
    rgb_array[:, :, 1] = g.reshape(height, width)
    rgb_array[:, :, 2] = b.reshape(height, width)
    
    # Convert to PIL Image and save as JPEG
    image = Image.fromarray(rgb_array, 'RGB')
    return image

def load_or_create_excel(file_name):
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()
        workbook.active.append(["Timestamp", "Ph_Value", "Turbidity", "Temperature", "Flow_Value"])  # Header
        workbook.save(file_name)
    return workbook

@app.route('/check', methods=['POST', 'GET', 'OPTIONS'])
def check():
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods','GET, POST, OPTIONS')
        return response
    
    print("\n" + "="*50)
    print("Received request from ESP32")
    print("="*50)
    
    try:
        # Log request details
        print(f"Content-Type: {request.content_type}")
        print(f"Content-Length: {request.content_length}")
        print(f"Form data keys: {list(request.form.keys()) if request.form else 'None'}")
        print(f"Files keys: {list(request.files.keys()) if request.files else 'None'}")
        
        # Get WiFi data for location if available
        wifi_data = request.form.get('wifi_data', None) if request.form else None
        latitude, longitude, location_info = None, None, "No location data"
        
        if wifi_data:
            print(f"WiFi data received: {len(wifi_data)} bytes")
            latitude, longitude, location_info = get_location_from_wifi(wifi_data)
            print(f"Location: {latitude}, {longitude} ({location_info})")
        else:
            print("No WiFi data provided")
        
        # Handle multipart form data from ESP32
        if 'photo' in request.files:
            print("Processing multipart form data...")
            
            photo = request.files['photo']
            if photo.filename == '':
                return jsonify({'error': 'No photo selected'}), 400
            
            print(f"Photo filename: {photo.filename}")
            print(f"Photo content type: {photo.content_type}")
            
            # Read the photo data
            photo_data = photo.read()
            print(f"Photo size: {len(photo_data)} bytes")
            
            # Generate timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            image_name = f"esp32_photo_{timestamp}.jpg"
            saved_photo_path = os.path.join(PHOTOS_FOLDER, image_name)
            
            # Save uploaded file temporarily
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
            temp_file_path = temp_file.name
            
            try:
                # Write photo data to temporary file
                with open(temp_file_path, 'wb') as f:
                    f.write(photo_data)
                
                # Save a copy to photos folder
                with open(saved_photo_path, 'wb') as f:
                    f.write(photo_data)
                    
                print(f"Photo saved: {saved_photo_path}")
                
                # Run inference
                print("Running Roboflow inference...")
                result = client.run_workflow(
                    workspace_name="microshets",
                    workflow_id="detect-and-classify-2",
                    images={
                        "image": temp_file_path
                    },
                    use_cache=True
                )
                
                print(f"Inference result: {result}")
                
                # Extract trash detection results
                trash_detected = False
                trash_name = "none"
                
                if result and len(result) > 0:
                    detection_predictions = result[0].get('detection_predictions', {})
                    predictions = detection_predictions.get('predictions', [])
                    
                    if predictions and len(predictions) > 0:
                        trash_detected = True
                        trash_name = predictions[0].get('class', 'unknown')
                        confidence = predictions[0].get('confidence', 0)
                        print(f"Trash detected: {trash_name} (confidence: {confidence:.2f})")
                    else:
                        print("No trash detected in image")
                else:
                    print("No predictions returned from model")
                
                # Log to Excel
                log_to_excel(image_name, latitude, longitude, location_info, 
                           trash_detected, trash_name, datetime.now())
                
            finally:
                # Clean up temporary file with error handling
                try:
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                except PermissionError:
                    pass  # File might still be in use, skip deletion
            
            # Return simplified results with CORS headers
            response_data = {
                'trash_detected': trash_detected,
                'result': trash_name,
                'location': {
                    'latitude': latitude,
                    'longitude': longitude,
                    'info': location_info
                },
                'image_saved': image_name,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            print(f"Sending response: {response_data}")
            print("="*50 + "\n")
            
            response = jsonify(response_data)
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        
        # Check if raw JPEG data was uploaded (alternative method)
        elif request.content_type and ('application/octet-stream' in request.content_type or 'image/jpeg' in request.content_type):
            print("Processing raw JPEG data...")
            
            # Handle direct JPEG data from ESP32
            jpeg_data = request.get_data()
            if len(jpeg_data) == 0:
                return jsonify({'error': 'No data received'}), 400
            
            print(f"JPEG data size: {len(jpeg_data)} bytes")
            
            # Generate timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            image_name = f"esp32_photo_{timestamp}.jpg"
            saved_photo_path = os.path.join(PHOTOS_FOLDER, image_name)
            
            # Save uploaded file temporarily
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
            temp_file_path = temp_file.name
            
            try:
                # Write JPEG data directly to file
                with open(temp_file_path, 'wb') as f:
                    f.write(jpeg_data)
                
                # Save a copy to photos folder
                with open(saved_photo_path, 'wb') as f:
                    f.write(jpeg_data)
                print(f"Photo saved: {saved_photo_path}")
                
                # Run inference
                print("Running Roboflow inference...")
                result = client.run_workflow(
                    workspace_name="microshets",
                    workflow_id="detect-and-classify-2",
                    images={
                        "image": temp_file_path
                    },
                    use_cache=True
                )
                
                # Extract trash detection results
                trash_detected = False
                trash_name = "none"
                
                if result and len(result) > 0:
                    detection_predictions = result[0].get('detection_predictions', {})
                    predictions = detection_predictions.get('predictions', [])
                    
                    if predictions and len(predictions) > 0:
                        trash_detected = True
                        trash_name = predictions[0].get('class', 'unknown')
                        print(f"Trash detected: {trash_name}")
                
                # Log to Excel
                log_to_excel(image_name, latitude, longitude, location_info, 
                           trash_detected, trash_name, datetime.now())
                
            finally:
                # Clean up temporary file with error handling
                try:
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                except PermissionError:
                    pass  # File might still be in use, skip deletion
            
            # Return simplified results with CORS headers
            response_data = {
                'trash_detected': trash_detected,
                'result': trash_name,
                'location': {
                    'latitude': latitude,
                    'longitude': longitude,
                    'info': location_info
                }
            }
            
            print(f"Sending response: {response_data}")
            print("="*50 + "\n")
            
            response = jsonify(response_data)
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
        
        else:
            print("ERROR: No photo data found in request")
            print("="*50 + "\n")
            return jsonify({'error': 'No photo uploaded. Please send photo as multipart/form-data with field name "photo"'}), 400
            
    except Exception as e:
        print(f"ERROR: {str(e)}")
        print("="*50 + "\n")
        import traceback
        traceback.print_exc()
        
        error_response = jsonify({'error': str(e)})
        error_response.headers.add('Access-Control-Allow-Origin', '*')
        return error_response, 500

# API route to handle POST requests and save data to Excel
@app.route('/add_to_excel', methods=['POST'])
def add_to_excel():
    data = request.get_json()

    # Automatically generate the timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ph_value = data.get('ph_value')
    turbidity = data.get('turbidity')
    temperature = data.get('temperature')
    flow_value = data.get('flow_value')

    if not all([ph_value, turbidity, temperature, flow_value]):
        return jsonify({"error": "Missing required parameters"}), 400

    # File name for the Excel file
    file_name = "data.xlsx"
    workbook = load_or_create_excel(file_name)
    sheet = workbook.active

    # Add the data to the Excel sheet
    sheet.append([timestamp, ph_value, turbidity, temperature, flow_value])
    workbook.save(file_name)

    return jsonify({"message": "Data added to Excel successfully!"}), 200

# API route to retrieve the latest row from the Excel file
@app.route('/get_entry', methods=['GET'])
def get_latest_entry():
    # File name for the Excel file
    file_name = "data.xlsx"
    
    if not os.path.exists(file_name):
        return jsonify({"error": "Excel file not found"}), 404

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    if sheet.max_row < 2:  # Check if there are any rows other than the header
        return jsonify({"error": "No data available"}), 404

    # Get the latest row (excluding the header)
    latest_row = list(sheet.iter_rows(values_only=True))[-1]
    
    # Create an object for the latest entry
    latest_entry = {
        "timestamp": latest_row[0],
        "ph_value": latest_row[1],
        "turbidity": latest_row[2],
        "temperature": latest_row[3],
        "flow_value": latest_row[4]
    }

    return jsonify({"latest_entry": latest_entry}), 200

@app.route('/predict_latest', methods=['GET'])
def predict_latest():
    file_name = "data.xlsx"
    if not os.path.exists(file_name):
        return jsonify({"error": "Excel file not found"}), 404

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    if sheet.max_row < 2:
        return jsonify({"error": "No data available"}), 404

    latest_row = list(sheet.iter_rows(values_only=True))[-1]
    features = {
        'temperature_C': latest_row[3],
        'pH': latest_row[1],
        'turbidity_NTU': latest_row[2],
        'flow_m_s': latest_row[4],
        'trash_detected': 0
    }

    try:
        prediction = predict(features)
        print(prediction)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"prediction": prediction}), 200

@app.route('/get_database', methods=['GET'])
def get_database():
    file_name = "data.xlsx"
    
    if not os.path.exists(file_name):
        return jsonify({"error": "Excel file not found"}), 404
    
    try:
        return send_file(file_name, as_attachment=True)
    except Exception as e:
        return str(e), 500

if __name__ == '__main__':
    # Start the Flask app on a specific port
    port = 5001

    print("\n" + "="*60)
    print("Starting Water Quality & Trash Detection Server")
    print("="*60)
    
    # Set your Ngrok auth token from environment
    ngrok_key = os.getenv("NGROK_AUTH_TOKEN")
    
    if ngrok_key:
        ngrok.set_auth_token(ngrok_key)
        
        # Open an Ngrok tunnel
        public_url = ngrok.connect(port).public_url
        print(f"\n✓ Ngrok tunnel active")
        print(f"  Public URL: {public_url}")
        print(f"  Local URL:  http://localhost:{port}")
        print(f"\nESP32 Configuration:")
        print(f"  const char* serverURL = \"{public_url}/check\";")
    else:
        print(f"\n✗ Ngrok not configured (no NGROK_AUTH_TOKEN in .env)")
        print(f"  Server running locally only")
        print(f"  Local URL: http://localhost:{port}")
        print(f"\nFor local network access, use your PC's IP address:")
        print(f"  Find your IP: ipconfig (Windows) or ifconfig (Linux/Mac)")
        print(f"  Example ESP32 config: const char* serverURL = \"http://192.168.1.100:{port}/check\";")
    
    print("\n" + "="*60)
    print("Available Endpoints:")
    print("="*60)
    print(f"  POST   /check              - Upload photo for trash detection")
    print(f"  POST   /add_to_excel       - Add water quality data")
    print(f"  GET    /get_entry          - Get latest water quality entry")
    print(f"  GET    /predict_latest     - ML prediction on latest data")
    print(f"  GET    /get_database       - Download Excel database")
    print("="*60 + "\n")
    
    print("Server is ready to receive ESP32 images!")
    print("Waiting for connections...\n")

    # Start the Flask server with external access
    app.run(host='0.0.0.0', port=port, debug=False)
